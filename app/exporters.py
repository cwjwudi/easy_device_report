from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, portrait
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen.canvas import Canvas as _PdfCanvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def html_escape(value: Any) -> str:
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def report_to_html(report: dict[str, Any]) -> str:
    def simple_table(rows: list[list[Any]]) -> str:
        return "<table>" + "".join("<tr>" + "".join(f"<td>{html_escape(cell)}</td>" for cell in row) + "</tr>" for row in rows) + "</table>"

    def query_table_html(columns: list[dict[str, Any]], rows: list[dict[str, Any]]) -> str:
        head = "<thead><tr>" + "".join(f"<th>{html_escape(c.get('label', c.get('name', '')))}</th>" for c in columns) + "</tr></thead>"
        body_rows = "".join(
            "<tr>" + "".join(f"<td>{html_escape(row.get(c.get('name', ''), ''))}</td>" for c in columns) + "</tr>"
            for row in rows
        )
        return f"<table>{head}<tbody>{body_rows}</tbody></table>"

    def body_tables_html(tables: list[dict[str, Any]]) -> str:
        chunks: list[str] = []
        for table in tables or []:
            title = f"<h2>{html_escape(table.get('title', ''))}</h2>" if table.get("title") else ""
            if table.get("kind") == "query":
                chunks.append(f"<section>{title}{query_table_html(table.get('columns', []), table.get('rows', []))}</section>")
            else:
                chunks.append(f"<section>{title}{simple_table(table.get('rows', []))}</section>")
        return "".join(chunks)

    body = report["body"]
    tables = body.get("tables")
    if not tables:
        # Legacy fallback: render custom tables then a single query table.
        legacy_custom = "".join(
            f"<section>{('<h2>' + html_escape(t.get('title', '')) + '</h2>') if t.get('title') else ''}{simple_table(t.get('rows', []))}</section>"
            for t in body.get("custom_tables", [])
        )
        legacy_query = query_table_html(body.get("columns", []), body.get("rows", []))
        body_html = legacy_custom + legacy_query
    else:
        body_html = body_tables_html(tables)
    return f"""
<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <title>{html_escape(report['name'])}</title>
  <style>
    body {{ font-family: Arial, "Microsoft YaHei", sans-serif; color: #1d2433; }}
    table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
    th, td {{ border: 1px solid #9aa5b1; padding: 7px 9px; font-size: 12px; }}
    th {{ background: #eef2f6; }}
    h1 {{ font-size: 20px; margin: 0 0 12px; }}
    h2 {{ font-size: 14px; margin: 12px 0 4px; }}
    .meta {{ color: #5c6676; font-size: 12px; }}
  </style>
</head>
<body>
  <h1>{html_escape(report['name'])}</h1>
  <div class="meta">Generated at {html_escape(report['generated_at'])}</div>
  {simple_table(report['header']['rows'])}
  {body_html}
  {simple_table(report['footer']['rows'])}
</body>
</html>
"""


def make_excel(report: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append([report["name"]])
    ws.append(["Generated At", report["generated_at"]])
    ws.append([])
    for row in report["header"]["rows"]:
        ws.append(row)
    ws.append([])

    def append_query(columns: list[dict[str, Any]], rows: list[dict[str, Any]]) -> None:
        if not columns:
            return
        ws.append([column.get("label", column.get("name", "")) for column in columns])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E8EEF6")
        for item in rows:
            ws.append([item.get(column.get("name", ""), "") for column in columns])
        ws.append([])

    tables = report["body"].get("tables")
    if tables:
        for table in tables:
            if table.get("title"):
                ws.append([table["title"]])
            if table.get("kind") == "query":
                append_query(table.get("columns", []), table.get("rows", []))
            else:
                for row in table.get("rows", []):
                    ws.append(row)
                ws.append([])
    else:
        for table in report["body"].get("custom_tables", []):
            if table.get("title"):
                ws.append([table["title"]])
            for row in table.get("rows", []):
                ws.append(row)
            ws.append([])
        append_query(report["body"].get("columns", []), report["body"].get("rows", []))
    for row in report["footer"]["rows"]:
        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top")
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 4
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width, 10), 36)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def make_pdf(report: dict[str, Any]) -> bytes:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    stream = io.BytesIO()
    page_size = landscape(A4) if report.get("page", {}).get("orientation") == "landscape" else portrait(A4)
    margin_mm = float(report.get("page", {}).get("margin_mm") or 14)
    header_repeat = bool(report.get("header", {}).get("repeat_pdf_each_page"))
    footer_repeat = bool(report.get("footer", {}).get("repeat_pdf_each_page"))
    page_number_position: str = report.get("page", {}).get("page_number_position") or "none"
    header_rows = report.get("header", {}).get("rows", [])
    footer_rows = report.get("footer", {}).get("rows", [])
    repeat_header_height = (len(header_rows) * 8 + 10) * mm if header_repeat and header_rows else 0
    repeat_footer_height = (len(footer_rows) * 8 + 10) * mm if footer_repeat and footer_rows else 0
    doc = SimpleDocTemplate(
        stream,
        pagesize=page_size,
        leftMargin=margin_mm * mm,
        rightMargin=margin_mm * mm,
        topMargin=(margin_mm * mm) + repeat_header_height,
        bottomMargin=(margin_mm * mm) + repeat_footer_height,
    )
    styles = getSampleStyleSheet()
    styles["Title"].fontName = "STSong-Light"
    styles["Normal"].fontName = "STSong-Light"
    elements: list[Any] = [Paragraph(report["name"], styles["Title"]), Paragraph(f"生成时间: {report['generated_at']}", styles["Normal"]), Spacer(1, 6)]

    def make_table(rows: list[list[Any]], header: bool = False) -> Table | None:
        if not rows:
            return None
        table = Table([[str(cell) for cell in row] for row in rows], repeatRows=1 if header else 0)
        style = [
            ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]
        if header:
            style += [("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF6"))]
        table.setStyle(TableStyle(style))
        return table

    def add_table(rows: list[list[Any]], header: bool = False) -> None:
        table = make_table(rows, header)
        if table is None:
            return
        elements.extend([table, Spacer(1, 8)])

    def draw_repeating_table(rows: list[list[Any]], canvas: Any, bottom_y: float) -> None:
        if not rows:
            return
        width = page_size[0] - doc.leftMargin - doc.rightMargin
        max_cols = max((len(row) for row in rows), default=1) or 1
        row_h = 8 * mm
        col_w = width / max_cols
        canvas.setFont("STSong-Light", 7)
        canvas.setStrokeColor(colors.grey)
        canvas.setLineWidth(0.4)
        for row_index, row in enumerate(rows):
            y = bottom_y + (len(rows) - row_index - 1) * row_h
            for col_index in range(max_cols):
                x = doc.leftMargin + col_index * col_w
                canvas.rect(x, y, col_w, row_h, stroke=1, fill=0)
                text = str(row[col_index] if col_index < len(row) else "")
                canvas.drawString(x + 2, y + row_h - 5 * mm, text[:42])

    def on_page(canvas: Any, _doc: Any) -> None:
        canvas.saveState()
        if header_repeat:
            header_h = len(header_rows) * 8 * mm
            draw_repeating_table(header_rows, canvas, page_size[1] - margin_mm * mm - header_h)
        if footer_repeat:
            draw_repeating_table(footer_rows, canvas, margin_mm * mm)
        canvas.restoreState()

    if not header_repeat:
        add_table(header_rows)

    def query_to_rows(columns: list[dict[str, Any]], data_rows: list[dict[str, Any]]) -> list[list[Any]]:
        if not columns:
            return []
        header_row = [column.get("label", column.get("name", "")) for column in columns]
        body = [[row.get(column.get("name", ""), "") for column in columns] for row in data_rows]
        return [header_row] + body

    tables = report["body"].get("tables")
    if tables:
        for table in tables:
            if table.get("title"):
                elements.append(Paragraph(str(table["title"]), styles["Normal"]))
            if table.get("kind") == "query":
                add_table(query_to_rows(table.get("columns", []), table.get("rows", [])), header=True)
            else:
                add_table(table.get("rows", []))
    else:
        for table in report["body"].get("custom_tables", []):
            if table.get("title"):
                elements.append(Paragraph(str(table["title"]), styles["Normal"]))
            add_table(table.get("rows", []))
        add_table(query_to_rows(report["body"].get("columns", []), report["body"].get("rows", [])), header=True)
    if not footer_repeat:
        add_table(footer_rows)

    # Two-pass canvas: draws page numbers only after total page count is known.
    class NumberedCanvas(_PdfCanvas):
        def __init__(self, *args: Any, **kwargs: Any) -> None:
            super().__init__(*args, **kwargs)
            self._page_states: list[dict] = []

        def showPage(self) -> None:  # type: ignore[override]
            self._page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self) -> None:  # type: ignore[override]
            total = len(self._page_states)
            for state in self._page_states:
                self.__dict__.update(state)
                if page_number_position != "none":
                    self._draw_page_num(total)
                _PdfCanvas.showPage(self)
            _PdfCanvas.save(self)

        def _draw_page_num(self, total: int) -> None:
            page_num = self._pageNumber
            text = f"{page_num} / {total}"
            self.saveState()
            self.setFont("STSong-Light", 8)
            self.setFillColor(colors.HexColor("#657184"))
            page_w, page_h = page_size
            usable_w = page_w - doc.leftMargin - doc.rightMargin
            text_w = self.stringWidth(text, "STSong-Light", 8)
            if "top" in page_number_position:
                y = page_h - margin_mm * mm * 0.65
            else:
                y = margin_mm * mm * 0.35
            if "left" in page_number_position:
                x = doc.leftMargin
            elif "right" in page_number_position:
                x = doc.leftMargin + usable_w - text_w
            else:
                x = doc.leftMargin + (usable_w - text_w) / 2
            self.drawString(x, y, text)
            self.restoreState()

    canvas_maker = NumberedCanvas if page_number_position != "none" else _PdfCanvas
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=canvas_maker)
    return stream.getvalue()
