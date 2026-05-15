from __future__ import annotations

import io
import json
import re
import sqlite3
from contextlib import contextmanager
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Literal

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, portrait
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BASE_DIR = Path(__file__).resolve().parent.parent
STATIC_DIR = BASE_DIR / "static"
DATA_DIR = BASE_DIR / "app_data"
APP_DB = DATA_DIR / "report_app.sqlite"
DEMO_DB = DATA_DIR / "business_data.sqlite"

IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
ALLOWED_OPERATORS = {"=", "!=", ">", ">=", "<", "<=", "LIKE"}

FIELD_OPCUA = {
    "server_url": "opc.tcp://192.168.50.233:4840",
    "root_node": "ns=6;i=1000",
}
FIELD_MYSQL = {
    "type": "mysql",
    "name": "wn_10",
    "host": "192.168.50.22",
    "port": 3306,
    "username": "root",
    "password": "Br54644800@",
    "database": "wn_10",
    "charset": "utf8mb4",
}


class OpcUaTestRequest(BaseModel):
    server_url: str = "mock://local"
    timeout_seconds: float = 5
    node_values: dict[str, Any] = Field(default_factory=dict)


class OpcUaReadRequest(OpcUaTestRequest):
    nodes: list[str] = Field(default_factory=list)


class OpcUaBrowseRequest(OpcUaTestRequest):
    root_node: str = "ns=6;i=1000"
    max_depth: int = 2
    limit: int = 120
    include_values: bool = True


class OpcUaPointPayload(BaseModel):
    alias: str
    node_id: str
    display_name: str | None = None
    browse_name: str | None = None
    server_url: str = FIELD_OPCUA["server_url"]
    root_node: str = FIELD_OPCUA["root_node"]
    data_type: str | None = None
    refresh_seconds: int = 5


class DatabaseConnection(BaseModel):
    type: Literal["sqlite", "mysql"] = "sqlite"
    path: str = str(DEMO_DB)
    name: str | None = None
    host: str = "127.0.0.1"
    port: int = 3306
    username: str = "root"
    password: str = ""
    database: str | None = None
    charset: str = "utf8mb4"


class QueryPreviewRequest(BaseModel):
    connection: DatabaseConnection = Field(default_factory=DatabaseConnection)
    table: str
    columns: list[Any] = Field(default_factory=list)
    filters: list[dict[str, Any]] = Field(default_factory=list)
    order_by: list[dict[str, Any]] = Field(default_factory=list)
    limit: int = 50
    opcua_values: dict[str, Any] = Field(default_factory=dict)


class TemplatePayload(BaseModel):
    name: str
    config: dict[str, Any]


class GenerateReportRequest(BaseModel):
    template_id: int | None = None
    template: dict[str, Any] | None = None
    persist_run: bool = True


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def make_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): make_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [make_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [make_jsonable(item) for item in value]
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def resolve_path(path_text: str) -> Path:
    path = Path(path_text)
    if not path.is_absolute():
        path = BASE_DIR / path
    return path


@contextmanager
def sqlite_connect(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def mysql_connect(connection: DatabaseConnection, include_database: bool = True):
    try:
        import pymysql
    except Exception as exc:  # pragma: no cover - dependency safety
        raise HTTPException(status_code=500, detail=f"PyMySQL is not available: {exc}") from exc

    kwargs = {
        "host": connection.host,
        "port": connection.port,
        "user": connection.username,
        "password": connection.password,
        "charset": connection.charset,
        "connect_timeout": 8,
        "read_timeout": 20,
        "write_timeout": 20,
        "cursorclass": pymysql.cursors.DictCursor,
    }
    if include_database and connection.database:
        kwargs["database"] = connection.database
    return pymysql.connect(**kwargs)


def validate_identifier(value: str, label: str = "identifier") -> str:
    if not IDENTIFIER_RE.match(value):
        raise HTTPException(status_code=400, detail=f"Invalid {label}: {value}")
    return value


def quote_identifier(value: str, database_type: str) -> str:
    validate_identifier(value)
    quote = "`" if database_type == "mysql" else '"'
    return f"{quote}{value}{quote}"


def normalize_columns(columns: list[Any]) -> list[dict[str, str]]:
    normalized = []
    for item in columns:
        if isinstance(item, str):
            normalized.append({"name": item, "label": item})
        else:
            name = str(item.get("name", ""))
            normalized.append({"name": name, "label": str(item.get("label") or name)})
    if not normalized:
        normalized = [{"name": "*", "label": "*"}]
    return normalized


async def read_opcua_values(config: dict[str, Any], nodes: list[str]) -> dict[str, Any]:
    server_url = config.get("server_url", "mock://local")
    node_values = config.get("node_values") or {}
    if server_url.startswith("mock://"):
        return {node: node_values.get(node, f"mock:{node}") for node in nodes}

    try:
        from asyncua import Client
    except Exception as exc:  # pragma: no cover - dependency safety
        raise HTTPException(status_code=500, detail=f"asyncua is not available: {exc}") from exc

    result: dict[str, Any] = {}
    try:
        async with Client(url=server_url) as client:
            for node in nodes:
                value = await client.get_node(node).read_value()
                result[node] = make_jsonable(value)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"OPC UA read failed: {exc}") from exc
    return result


async def browse_opcua_nodes(request: OpcUaBrowseRequest) -> dict[str, Any]:
    if request.server_url.startswith("mock://"):
        nodes = [
            {"node_id": "batch_no", "display_name": "batch_no", "browse_name": "batch_no", "depth": 0, "value": "B20260514"},
            {"node_id": "shift", "display_name": "shift", "browse_name": "shift", "depth": 0, "value": "A"},
        ]
        return {"ok": True, "mode": "mock", "nodes": nodes}

    try:
        from asyncua import Client
    except Exception as exc:  # pragma: no cover - dependency safety
        raise HTTPException(status_code=500, detail=f"asyncua is not available: {exc}") from exc

    direct_children_only = request.max_depth == 0
    max_depth = max(0, min(request.max_depth, 6))
    limit = max(1, min(request.limit, 500))
    results: list[dict[str, Any]] = []
    try:
        async with Client(url=request.server_url) as client:
            root_node = client.get_node(request.root_node)
            queue = []
            if direct_children_only:
                for child in await root_node.get_children():
                    queue.append((child, 0))
                if not queue:
                    queue = [(root_node, 0)]
            else:
                queue = [(root_node, 0)]
            seen: set[str] = set()
            while queue and len(results) < limit:
                node, depth = queue.pop(0)
                node_id = node.nodeid.to_string()
                if node_id in seen:
                    continue
                seen.add(node_id)
                display = await node.read_display_name()
                browse = await node.read_browse_name()
                item = {
                    "node_id": node_id,
                    "display_name": display.Text,
                    "browse_name": browse.Name,
                    "depth": depth,
                }
                try:
                    children = await node.get_children()
                    item["has_children"] = len(children) > 0
                except Exception:
                    children = []
                    item["has_children"] = False
                if request.include_values:
                    try:
                        item["value"] = make_jsonable(await node.read_value())
                    except Exception:
                        item["value"] = None
                if not (direct_children_only and node_id == request.root_node and results):
                    results.append(item)
                if not direct_children_only and depth < max_depth:
                    for child in children:
                        queue.append((child, depth + 1))
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"OPC UA browse failed: {exc}") from exc
    return {"ok": True, "mode": "real", "nodes": results}


async def test_opcua_connection(request: OpcUaTestRequest) -> dict[str, Any]:
    if request.server_url.startswith("mock://"):
        return {"ok": True, "mode": "mock", "message": "Mock OPC UA connection is available."}
    await read_opcua_values(request.model_dump(), [])
    return {"ok": True, "mode": "real", "message": "OPC UA connection succeeded."}


def build_select_query(request: QueryPreviewRequest) -> tuple[str, list[Any], list[dict[str, str]]]:
    database_type = request.connection.type
    placeholder = "%s" if database_type == "mysql" else "?"
    table = validate_identifier(request.table, "table")
    columns = normalize_columns(request.columns)
    sql_columns: list[str] = []
    for column in columns:
        name = column["name"]
        if name == "*":
            sql_columns.append("*")
        else:
            sql_columns.append(quote_identifier(name, database_type))

    params: list[Any] = []
    where_parts: list[str] = []
    for filter_item in request.filters:
        column = validate_identifier(str(filter_item.get("column", "")), "filter column")
        operator = str(filter_item.get("operator", "=")).upper()
        if operator not in ALLOWED_OPERATORS:
            raise HTTPException(status_code=400, detail=f"Unsupported operator: {operator}")
        source = filter_item.get("source") or {"type": "literal", "value": filter_item.get("value")}
        if source.get("type") == "opcua":
            value = request.opcua_values.get(str(source.get("node_id", "")))
        else:
            value = source.get("value")
        where_parts.append(f"{quote_identifier(column, database_type)} {operator} {placeholder}")
        params.append(value)

    order_parts: list[str] = []
    for order_item in request.order_by:
        column = validate_identifier(str(order_item.get("column", "")), "order column")
        direction = str(order_item.get("direction", "ASC")).upper()
        if direction not in {"ASC", "DESC"}:
            raise HTTPException(status_code=400, detail=f"Unsupported order direction: {direction}")
        order_parts.append(f"{quote_identifier(column, database_type)} {direction}")

    if database_type == "mysql" and request.connection.database:
        table_ref = f"{quote_identifier(request.connection.database, database_type)}.{quote_identifier(table, database_type)}"
    else:
        table_ref = quote_identifier(table, database_type)

    limit = max(1, min(int(request.limit), 1000))
    sql = f"SELECT {', '.join(sql_columns)} FROM {table_ref}"
    if where_parts:
        sql += " WHERE " + " AND ".join(where_parts)
    if order_parts:
        sql += " ORDER BY " + ", ".join(order_parts)
    sql += f" LIMIT {limit}"
    return sql, params, columns


def run_query(request: QueryPreviewRequest) -> dict[str, Any]:
    sql, params, columns = build_select_query(request)
    if request.connection.type == "sqlite":
        path = resolve_path(request.connection.path)
        if not path.exists():
            raise HTTPException(status_code=404, detail=f"Database not found: {path}")
        try:
            with sqlite_connect(path) as conn:
                rows = [dict(row) for row in conn.execute(sql, params).fetchall()]
        except sqlite3.Error as exc:
            raise HTTPException(status_code=400, detail=f"Database query failed: {exc}") from exc
    elif request.connection.type == "mysql":
        try:
            conn = mysql_connect(request.connection)
            try:
                with conn.cursor() as cursor:
                    cursor.execute(sql, params)
                    rows = list(cursor.fetchall())
            finally:
                conn.close()
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"MySQL query failed: {exc}") from exc
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported database type: {request.connection.type}")

    rows = make_jsonable(rows)
    if columns and columns[0]["name"] == "*" and rows:
        columns = [{"name": key, "label": key} for key in rows[0].keys()]
    return {"sql": sql, "params": params, "columns": columns, "rows": rows, "row_count": len(rows)}


def inspect_database_schema(connection: DatabaseConnection, table_limit: int = 50) -> dict[str, Any]:
    if connection.type == "sqlite":
        path = resolve_path(connection.path)
        if not path.exists():
            raise HTTPException(status_code=404, detail=f"Database not found: {path}")
        with sqlite_connect(path) as conn:
            tables = [row["name"] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name LIMIT ?", (table_limit,)).fetchall()]
            result_tables = []
            for table in tables:
                columns = [dict(row) for row in conn.execute(f"PRAGMA table_info({quote_identifier(table, 'sqlite')})").fetchall()]
                result_tables.append({"database": None, "table": table, "columns": columns})
        return {"type": "sqlite", "tables": result_tables}

    conn = mysql_connect(connection, include_database=False)
    try:
        with conn.cursor() as cursor:
            if connection.database:
                databases = [connection.database]
            else:
                cursor.execute("SHOW DATABASES")
                databases = [
                    row["Database"]
                    for row in cursor.fetchall()
                    if row["Database"] not in {"information_schema", "mysql", "performance_schema", "sys"}
                ]
            result_tables = []
            for database in databases:
                validate_identifier(database, "database")
                cursor.execute(f"SHOW FULL TABLES FROM `{database}` WHERE Table_type = 'BASE TABLE'")
                table_rows = cursor.fetchall()
                for table_row in table_rows[:table_limit]:
                    table = next(value for key, value in table_row.items() if key.startswith("Tables_in_"))
                    cursor.execute(f"SHOW COLUMNS FROM `{database}`.`{table}`")
                    result_tables.append({"database": database, "table": table, "columns": make_jsonable(cursor.fetchall())})
    finally:
        conn.close()
    return {"type": "mysql", "tables": result_tables}


def get_template(template_id: int) -> dict[str, Any]:
    with sqlite_connect(APP_DB) as conn:
        row = conn.execute("SELECT id, name, config_json FROM report_templates WHERE id = ?", (template_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Report template not found.")
    config = json.loads(row["config_json"])
    config["id"] = row["id"]
    config["name"] = row["name"]
    return config


def normalize_body_tables(body: dict[str, Any]) -> list[dict[str, Any]]:
    """Return a normalized list of body tables (mix of query/custom items).

    Backwards compatible with the legacy layout where ``body`` holds a single
    query (table/columns/filters/order_by/limit) plus a ``custom_tables`` list.
    """

    raw = body.get("tables")
    if isinstance(raw, list) and raw:
        result: list[dict[str, Any]] = []
        for index, item in enumerate(raw):
            if not isinstance(item, dict):
                continue
            kind = item.get("kind")
            entry: dict[str, Any] = {
                "id": str(item.get("id") or f"t_{index}"),
                "title": item.get("title") or "",
            }
            if kind == "custom":
                entry["kind"] = "custom"
                entry["rows"] = item.get("rows", [])
            else:
                entry["kind"] = "query"
                entry["table"] = item.get("table", "")
                entry["columns"] = item.get("columns", [])
                entry["filters"] = item.get("filters", [])
                entry["order_by"] = item.get("order_by", [])
                entry["limit"] = item.get("limit", 500)
            result.append(entry)
        return result

    # Legacy layout: synthesize a tables list.
    legacy: list[dict[str, Any]] = []
    has_query = any(body.get(key) for key in ("table", "columns", "filters", "order_by")) or "limit" in body
    if has_query:
        legacy.append(
            {
                "id": "t_query_legacy",
                "kind": "query",
                "title": "",
                "table": body.get("table", ""),
                "columns": body.get("columns", []),
                "filters": body.get("filters", []),
                "order_by": body.get("order_by", []),
                "limit": body.get("limit", 500),
            }
        )
    for index, table in enumerate(body.get("custom_tables", []) or []):
        if not isinstance(table, dict):
            continue
        legacy.append(
            {
                "id": f"t_custom_legacy_{index}",
                "kind": "custom",
                "title": table.get("title", ""),
                "rows": table.get("rows", []),
            }
        )
    return legacy


def collect_opcua_nodes(template: dict[str, Any]) -> list[str]:
    nodes: set[str] = set()
    for region_name in ("header", "footer"):
        for row in template.get(region_name, {}).get("rows", []):
            for cell in row:
                if isinstance(cell, dict) and cell.get("type") == "opcua" and cell.get("node_id"):
                    nodes.add(str(cell["node_id"]))
    body = template.get("body", {}) or {}
    for item in normalize_body_tables(body):
        if item["kind"] == "custom":
            for row in item.get("rows", []):
                for cell in row:
                    if isinstance(cell, dict) and cell.get("type") == "opcua" and cell.get("node_id"):
                        nodes.add(str(cell["node_id"]))
        else:
            for filter_item in item.get("filters", []) or []:
                source = (filter_item or {}).get("source") or {}
                if source.get("type") == "opcua" and source.get("node_id"):
                    nodes.add(str(source["node_id"]))
    return sorted(nodes)


def resolve_cell(
    cell: Any,
    opcua_values: dict[str, Any],
    body_rows: list[dict[str, Any]],
    rows_by_source: dict[str, list[dict[str, Any]]] | None = None,
) -> Any:
    if not isinstance(cell, dict):
        return cell
    cell_type = cell.get("type", "static")
    if cell_type == "static":
        return cell.get("value", "")
    if cell_type == "opcua":
        return opcua_values.get(str(cell.get("node_id", "")), "")
    source_id = cell.get("source_id")
    rows = body_rows
    if source_id and rows_by_source and source_id in rows_by_source:
        rows = rows_by_source[source_id]
    if cell_type == "db_summary":
        aggregate = cell.get("aggregate", "count")
        column = cell.get("column")
        if aggregate == "count":
            return len(rows)
        if aggregate == "sum" and column:
            return round(sum(float(row.get(column) or 0) for row in rows), 4)
        if aggregate == "avg" and column and rows:
            values = [float(row.get(column) or 0) for row in rows]
            return round(sum(values) / len(values), 4)
    if cell_type == "db_field" and rows:
        return rows[0].get(str(cell.get("column", "")), "")
    return ""


def render_region(
    region: dict[str, Any],
    opcua_values: dict[str, Any],
    body_rows: list[dict[str, Any]],
    rows_by_source: dict[str, list[dict[str, Any]]] | None = None,
) -> dict[str, Any]:
    rows = []
    for row in region.get("rows", []):
        rows.append([resolve_cell(cell, opcua_values, body_rows, rows_by_source) for cell in row])
    return {"title": region.get("title", ""), "rows": rows}


async def generate_report(template: dict[str, Any]) -> dict[str, Any]:
    opcua_config = template.get("opcua", {})
    nodes = collect_opcua_nodes(template)
    opcua_values = await read_opcua_values(opcua_config, nodes)

    body = template.get("body", {}) or {}
    body_items = normalize_body_tables(body)

    # Execute each query item; collect rows keyed by item id (for db_field/db_summary lookup).
    rows_by_source: dict[str, list[dict[str, Any]]] = {}
    query_outputs: dict[str, dict[str, Any]] = {}
    first_query_rows: list[dict[str, Any]] = []
    connection_payload = template.get("database", {}) or {}

    for item in body_items:
        if item["kind"] != "query":
            continue
        query_request = QueryPreviewRequest(
            connection=DatabaseConnection(**connection_payload),
            table=item.get("table") or "production_records",
            columns=[column.get("name") if isinstance(column, dict) else column for column in item.get("columns", [])],
            filters=item.get("filters", []),
            order_by=item.get("order_by", []),
            limit=int(item.get("limit", 500) or 500),
            opcua_values=opcua_values,
        )
        result = run_query(query_request)
        label_map = {column["name"]: column["label"] for column in normalize_columns(item.get("columns", []))}
        result_columns = [
            {"name": c["name"], "label": label_map.get(c["name"], c["label"])} for c in result["columns"]
        ]
        query_outputs[item["id"]] = {
            "columns": result_columns,
            "rows": result["rows"],
            "sql": result["sql"],
            "row_count": result["row_count"],
        }
        rows_by_source[item["id"]] = result["rows"]
        if not first_query_rows:
            first_query_rows = result["rows"]

    # Build the ordered output table list (mirrors template order).
    out_tables: list[dict[str, Any]] = []
    legacy_custom: list[dict[str, Any]] = []
    for item in body_items:
        if item["kind"] == "query":
            q = query_outputs.get(item["id"], {"columns": [], "rows": [], "sql": "", "row_count": 0})
            out_tables.append(
                {
                    "id": item["id"],
                    "kind": "query",
                    "title": item.get("title", ""),
                    "table": item.get("table", ""),
                    "columns": q["columns"],
                    "rows": q["rows"],
                    "sql": q["sql"],
                    "row_count": q["row_count"],
                }
            )
        else:
            rendered = render_region(item, opcua_values, first_query_rows, rows_by_source)
            entry = {
                "id": item["id"],
                "kind": "custom",
                "title": item.get("title", ""),
                "rows": rendered["rows"],
            }
            out_tables.append(entry)
            legacy_custom.append({"title": entry["title"], "rows": entry["rows"]})

    # Legacy fields: first query result populates body.columns/rows/sql/row_count.
    first_query_output = next((t for t in out_tables if t["kind"] == "query"), None)
    legacy_columns = first_query_output["columns"] if first_query_output else []
    legacy_rows = first_query_output["rows"] if first_query_output else []
    legacy_sql = first_query_output["sql"] if first_query_output else ""
    legacy_row_count = first_query_output["row_count"] if first_query_output else 0

    report = {
        "template_id": template.get("id"),
        "name": template.get("name", "Untitled Report"),
        "generated_at": now_iso(),
        "page": template.get("page", {}),
        "opcua_values": opcua_values,
        "header": render_region(template.get("header", {}), opcua_values, first_query_rows, rows_by_source),
        "body": {
            "tables": out_tables,
            "columns": legacy_columns,
            "rows": legacy_rows,
            "sql": legacy_sql,
            "row_count": legacy_row_count,
            "custom_tables": legacy_custom,
        },
        "footer": render_region(template.get("footer", {}), opcua_values, first_query_rows, rows_by_source),
    }
    return make_jsonable(report)


def html_escape(value: Any) -> str:
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def report_to_html(report: dict[str, Any]) -> str:
    def simple_table(rows: list[list[Any]]) -> str:
        return "<table>" + "".join("<tr>" + "".join(f"<td>{html_escape(cell)}</td>" for cell in row) + "</tr>" for row in rows) + "</table>"

    def query_table_html(columns: list[dict[str, Any]], rows: list[dict[str, Any]]) -> str:
        head = "<thead><tr>" + "".join(f"<th>{html_escape(c.get('label', c.get('name', '')))}</th>" for c in columns) + "</tr></thead>"
        body_rows = "".join(
            "<tr>" + "".join(f"<td>{html_escape(row.get(c.get('name', ''), ''))}</td>" for c in columns) + "</tr>"
            for row in rows
        )
        return f"<table>{head}<tbody>{body_rows}</tbody></table>"

    def body_tables_html(tables: list[dict[str, Any]]) -> str:
        chunks: list[str] = []
        for table in tables or []:
            title = f"<h2>{html_escape(table.get('title', ''))}</h2>" if table.get("title") else ""
            if table.get("kind") == "query":
                chunks.append(f"<section>{title}{query_table_html(table.get('columns', []), table.get('rows', []))}</section>")
            else:
                chunks.append(f"<section>{title}{simple_table(table.get('rows', []))}</section>")
        return "".join(chunks)

    body = report["body"]
    tables = body.get("tables")
    if not tables:
        # Legacy fallback: render custom tables then a single query table.
        legacy_custom = "".join(
            f"<section>{('<h2>' + html_escape(t.get('title', '')) + '</h2>') if t.get('title') else ''}{simple_table(t.get('rows', []))}</section>"
            for t in body.get("custom_tables", [])
        )
        legacy_query = query_table_html(body.get("columns", []), body.get("rows", []))
        body_html = legacy_custom + legacy_query
    else:
        body_html = body_tables_html(tables)
    return f"""
<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <title>{html_escape(report['name'])}</title>
  <style>
    body {{ font-family: Arial, "Microsoft YaHei", sans-serif; color: #1d2433; }}
    table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
    th, td {{ border: 1px solid #9aa5b1; padding: 7px 9px; font-size: 12px; }}
    th {{ background: #eef2f6; }}
    h1 {{ font-size: 20px; margin: 0 0 12px; }}
    h2 {{ font-size: 14px; margin: 12px 0 4px; }}
    .meta {{ color: #5c6676; font-size: 12px; }}
  </style>
</head>
<body>
  <h1>{html_escape(report['name'])}</h1>
  <div class="meta">Generated at {html_escape(report['generated_at'])}</div>
  {simple_table(report['header']['rows'])}
  {body_html}
  {simple_table(report['footer']['rows'])}
</body>
</html>
"""


def make_excel(report: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append([report["name"]])
    ws.append(["Generated At", report["generated_at"]])
    ws.append([])
    for row in report["header"]["rows"]:
        ws.append(row)
    ws.append([])

    def append_query(columns: list[dict[str, Any]], rows: list[dict[str, Any]]) -> None:
        if not columns:
            return
        ws.append([column.get("label", column.get("name", "")) for column in columns])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E8EEF6")
        for item in rows:
            ws.append([item.get(column.get("name", ""), "") for column in columns])
        ws.append([])

    tables = report["body"].get("tables")
    if tables:
        for table in tables:
            if table.get("title"):
                ws.append([table["title"]])
            if table.get("kind") == "query":
                append_query(table.get("columns", []), table.get("rows", []))
            else:
                for row in table.get("rows", []):
                    ws.append(row)
                ws.append([])
    else:
        for table in report["body"].get("custom_tables", []):
            if table.get("title"):
                ws.append([table["title"]])
            for row in table.get("rows", []):
                ws.append(row)
            ws.append([])
        append_query(report["body"].get("columns", []), report["body"].get("rows", []))
    for row in report["footer"]["rows"]:
        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top")
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 4
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width, 10), 36)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def make_pdf(report: dict[str, Any]) -> bytes:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    stream = io.BytesIO()
    page_size = landscape(A4) if report.get("page", {}).get("orientation") == "landscape" else portrait(A4)
    doc = SimpleDocTemplate(stream, pagesize=page_size, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = "STSong-Light"
    styles["Normal"].fontName = "STSong-Light"
    elements: list[Any] = [Paragraph(report["name"], styles["Title"]), Paragraph(f"生成时间: {report['generated_at']}", styles["Normal"]), Spacer(1, 6)]

    def add_table(rows: list[list[Any]], header: bool = False) -> None:
        if not rows:
            return
        table = Table([[str(cell) for cell in row] for row in rows], repeatRows=1 if header else 0)
        style = [
            ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]
        if header:
            style += [("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF6"))]
        table.setStyle(TableStyle(style))
        elements.extend([table, Spacer(1, 8)])

    add_table(report["header"]["rows"])

    def query_to_rows(columns: list[dict[str, Any]], data_rows: list[dict[str, Any]]) -> list[list[Any]]:
        if not columns:
            return []
        header_row = [column.get("label", column.get("name", "")) for column in columns]
        body = [[row.get(column.get("name", ""), "") for column in columns] for row in data_rows]
        return [header_row] + body

    tables = report["body"].get("tables")
    if tables:
        for table in tables:
            if table.get("title"):
                elements.append(Paragraph(str(table["title"]), styles["Normal"]))
            if table.get("kind") == "query":
                add_table(query_to_rows(table.get("columns", []), table.get("rows", [])), header=True)
            else:
                add_table(table.get("rows", []))
    else:
        for table in report["body"].get("custom_tables", []):
            if table.get("title"):
                elements.append(Paragraph(str(table["title"]), styles["Normal"]))
            add_table(table.get("rows", []))
        add_table(query_to_rows(report["body"].get("columns", []), report["body"].get("rows", [])), header=True)
    add_table(report["footer"]["rows"])
    doc.build(elements)
    return stream.getvalue()


def init_app_db() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with sqlite_connect(APP_DB) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS data_sources (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                kind TEXT NOT NULL,
                config_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS report_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                config_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS report_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                template_id INTEGER,
                status TEXT NOT NULL,
                message TEXT,
                rendered_json TEXT,
                created_at TEXT NOT NULL
            )
            """
        )


def init_demo_db() -> None:
    with sqlite_connect(DEMO_DB) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS production_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_no TEXT NOT NULL,
                record_time TEXT NOT NULL,
                line_name TEXT NOT NULL,
                product_code TEXT NOT NULL,
                temperature REAL NOT NULL,
                pressure REAL NOT NULL,
                quantity INTEGER NOT NULL,
                status TEXT NOT NULL
            )
            """
        )
        count = conn.execute("SELECT COUNT(*) AS value FROM production_records").fetchone()["value"]
        if count == 0:
            rows = [
                ("B20260514", "2026-05-14 08:00", "Line-1", "P-1001", 71.2, 0.83, 120, "OK"),
                ("B20260514", "2026-05-14 09:00", "Line-1", "P-1001", 72.1, 0.84, 118, "OK"),
                ("B20260514", "2026-05-14 10:00", "Line-1", "P-1001", 73.0, 0.86, 121, "OK"),
                ("B20260513", "2026-05-13 08:00", "Line-2", "P-1002", 69.8, 0.80, 110, "OK"),
            ]
            conn.executemany(
                """
                INSERT INTO production_records
                (batch_no, record_time, line_name, product_code, temperature, pressure, quantity, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                rows,
            )


def sample_sqlite_template() -> dict[str, Any]:
    return {
        "name": "生产批次报表",
        "page": {"size": "A4", "orientation": "portrait", "margin_mm": 14},
        "opcua": {"server_url": "mock://local", "node_values": {"batch_no": "B20260514", "shift": "A", "operator": "张三"}},
        "database": {"type": "sqlite", "path": str(DEMO_DB)},
        "header": {
            "title": "页眉",
            "rows": [
                [{"type": "static", "value": "批次号"}, {"type": "opcua", "node_id": "batch_no"}, {"type": "static", "value": "班次"}, {"type": "opcua", "node_id": "shift"}],
                [{"type": "static", "value": "操作员"}, {"type": "opcua", "node_id": "operator"}, {"type": "static", "value": "数据来源"}, {"type": "static", "value": "OPC UA + SQLite"}],
            ],
        },
        "body": {
            "table": "production_records",
            "columns": [
                {"name": "record_time", "label": "时间"},
                {"name": "line_name", "label": "产线"},
                {"name": "product_code", "label": "产品"},
                {"name": "temperature", "label": "温度"},
                {"name": "pressure", "label": "压力"},
                {"name": "quantity", "label": "数量"},
                {"name": "status", "label": "状态"},
            ],
            "filters": [{"column": "batch_no", "operator": "=", "source": {"type": "opcua", "node_id": "batch_no"}}],
            "order_by": [{"column": "record_time", "direction": "ASC"}],
            "limit": 500,
        },
        "footer": {
            "title": "页脚",
            "rows": [
                [{"type": "static", "value": "记录数"}, {"type": "db_summary", "aggregate": "count"}, {"type": "static", "value": "总数量"}, {"type": "db_summary", "aggregate": "sum", "column": "quantity"}],
                [{"type": "static", "value": "审核"}, {"type": "static", "value": ""}, {"type": "static", "value": "备注"}, {"type": "static", "value": ""}],
            ],
        },
    }


def field_mysql_template() -> dict[str, Any]:
    return {
        "name": "现场 MySQL + OPC UA 示例报表",
        "page": {"size": "A4", "orientation": "landscape", "margin_mm": 12},
        "opcua": FIELD_OPCUA,
        "database": FIELD_MYSQL,
        "header": {
            "title": "页眉",
            "rows": [
                [{"type": "static", "value": "OPC UA 服务器"}, {"type": "static", "value": FIELD_OPCUA["server_url"]}, {"type": "static", "value": "MySQL库"}, {"type": "static", "value": "wn_10"}],
                [{"type": "static", "value": "AIR实时值"}, {"type": "opcua", "node_id": "ns=6;s=::DataGen:AIR"}, {"type": "static", "value": "AP1实时值"}, {"type": "opcua", "node_id": "ns=6;s=::DataGen:AP1"}],
                [{"type": "static", "value": "BP1实时值"}, {"type": "opcua", "node_id": "ns=6;s=::DataGen:BP1"}, {"type": "static", "value": "DB筛选"}, {"type": "static", "value": "AIR >= OPC UA AIR"}],
            ],
        },
        "body": {
            "table": "Group1_20260511",
            "columns": [
                {"name": "id", "label": "ID"},
                {"name": "collection_time", "label": "采集时间"},
                {"name": "AIR", "label": "AIR"},
                {"name": "AP1", "label": "AP1"},
                {"name": "AP2", "label": "AP2"},
                {"name": "CP", "label": "CP"},
                {"name": "CT", "label": "CT"},
                {"name": "EC", "label": "EC"},
                {"name": "ES", "label": "ES"},
            ],
            "filters": [{"column": "AIR", "operator": ">=", "source": {"type": "opcua", "node_id": "ns=6;s=::DataGen:AIR"}}],
            "order_by": [{"column": "collection_time", "direction": "ASC"}],
            "limit": 80,
        },
        "footer": {
            "title": "页脚",
            "rows": [
                [{"type": "static", "value": "记录数"}, {"type": "db_summary", "aggregate": "count"}, {"type": "static", "value": "AIR平均值"}, {"type": "db_summary", "aggregate": "avg", "column": "AIR"}],
                [{"type": "static", "value": "CP合计"}, {"type": "db_summary", "aggregate": "sum", "column": "CP"}, {"type": "static", "value": "CT平均值"}, {"type": "db_summary", "aggregate": "avg", "column": "CT"}],
            ],
        },
    }


def upsert_template(conn: sqlite3.Connection, template: dict[str, Any]) -> None:
    row = conn.execute("SELECT id FROM report_templates WHERE name = ?", (template["name"],)).fetchone()
    payload = json.dumps(template, ensure_ascii=False)
    if row:
        conn.execute("UPDATE report_templates SET config_json = ?, updated_at = ? WHERE id = ?", (payload, now_iso(), row["id"]))
    else:
        conn.execute("INSERT INTO report_templates (name, config_json, created_at, updated_at) VALUES (?, ?, ?, ?)", (template["name"], payload, now_iso(), now_iso()))


def upsert_data_source(conn: sqlite3.Connection, name: str, kind: str, config: dict[str, Any]) -> None:
    row = conn.execute("SELECT id FROM data_sources WHERE name = ?", (name,)).fetchone()
    payload = json.dumps(config, ensure_ascii=False)
    if row:
        conn.execute("UPDATE data_sources SET kind = ?, config_json = ?, updated_at = ? WHERE id = ?", (kind, payload, now_iso(), row["id"]))
    else:
        conn.execute("INSERT INTO data_sources (name, kind, config_json, created_at, updated_at) VALUES (?, ?, ?, ?, ?)", (name, kind, payload, now_iso(), now_iso()))


def seed_defaults() -> None:
    with sqlite_connect(APP_DB) as conn:
        if conn.execute("SELECT COUNT(*) AS value FROM report_templates").fetchone()["value"] == 0:
            upsert_template(conn, sample_sqlite_template())
        upsert_template(conn, field_mysql_template())
        upsert_data_source(conn, "演示 SQLite 数据库", "database", {"type": "sqlite", "path": str(DEMO_DB)})
        upsert_data_source(conn, "现场 MySQL wn_10", "database", FIELD_MYSQL)
        upsert_data_source(conn, "现场 OPC UA", "opcua", FIELD_OPCUA)


def row_to_data_source(row: sqlite3.Row) -> dict[str, Any]:
    item = dict(row)
    item["config"] = json.loads(item["config_json"])
    item.pop("config_json", None)
    return item


app = FastAPI(title="Report Generator", version="0.2.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.on_event("startup")
def on_startup() -> None:
    init_app_db()
    init_demo_db()
    seed_defaults()


@app.get("/")
def index() -> FileResponse:
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/api/health")
def health() -> dict[str, Any]:
    return {"ok": True, "time": now_iso(), "app_db": str(APP_DB), "demo_db": str(DEMO_DB), "field_opcua": FIELD_OPCUA, "field_mysql": {**FIELD_MYSQL, "password": "***"}}


@app.post("/api/startup/one-click")
async def one_click_startup() -> dict[str, Any]:
    steps: list[dict[str, Any]] = []

    def record(name: str, ok: bool, detail: Any = None) -> None:
        steps.append({"name": name, "ok": ok, "detail": make_jsonable(detail)})

    try:
        init_app_db()
        init_demo_db()
        seed_defaults()
        record("初始化本地配置库和演示数据", True, {"app_db": str(APP_DB), "demo_db": str(DEMO_DB)})
    except Exception as exc:
        record("初始化本地配置库和演示数据", False, str(exc))

    try:
        opcua_result = await test_opcua_connection(OpcUaTestRequest(**FIELD_OPCUA))
        record("连接 OPC UA 服务器", True, opcua_result)
    except Exception as exc:
        record("连接 OPC UA 服务器", False, str(getattr(exc, "detail", exc)))

    try:
        nodes = ["ns=6;s=::DataGen:AIR", "ns=6;s=::DataGen:AP1", "ns=6;s=::DataGen:BP1"]
        values = await read_opcua_values(FIELD_OPCUA, nodes)
        record("读取 OPC UA 参考节点", True, values)
    except Exception as exc:
        record("读取 OPC UA 参考节点", False, str(getattr(exc, "detail", exc)))

    try:
        mysql_conn = DatabaseConnection(**FIELD_MYSQL)
        conn = mysql_connect(mysql_conn)
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT 1 AS ok")
                row = cursor.fetchone()
        finally:
            conn.close()
        record("连接 MySQL 数据库", True, row)
    except Exception as exc:
        record("连接 MySQL 数据库", False, str(getattr(exc, "detail", exc)))

    try:
        schema = inspect_database_schema(DatabaseConnection(**FIELD_MYSQL), table_limit=30)
        target = next((item for item in schema["tables"] if item["table"] == "Group1_20260511"), schema["tables"][0] if schema["tables"] else None)
        record("读取 MySQL 表结构", True, {"table_count": len(schema["tables"]), "target_table": target})
    except Exception as exc:
        record("读取 MySQL 表结构", False, str(getattr(exc, "detail", exc)))

    report: dict[str, Any] | None = None
    template_id: int | None = None
    try:
        with sqlite_connect(APP_DB) as conn:
            row = conn.execute("SELECT id FROM report_templates WHERE name = ?", (field_mysql_template()["name"],)).fetchone()
            template_id = row["id"] if row else None
        if template_id is None:
            raise RuntimeError("现场 MySQL + OPC UA 示例报表模板不存在")
        report = await generate_report(get_template(template_id))
        record("生成现场示例报表", True, {"template_id": template_id, "rows": report["body"]["row_count"], "sql": report["body"]["sql"]})
    except Exception as exc:
        record("生成现场示例报表", False, str(getattr(exc, "detail", exc)))

    ok = all(step["ok"] for step in steps)
    return {
        "ok": ok,
        "time": now_iso(),
        "template_id": template_id,
        "report": report,
        "steps": steps,
    }


@app.post("/api/opcua/test")
async def opcua_test(request: OpcUaTestRequest) -> dict[str, Any]:
    return await test_opcua_connection(request)


@app.post("/api/opcua/read")
async def opcua_read(request: OpcUaReadRequest) -> dict[str, Any]:
    values = await read_opcua_values(request.model_dump(), request.nodes)
    return {"ok": True, "values": values}


@app.post("/api/opcua/browse")
async def opcua_browse(request: OpcUaBrowseRequest) -> dict[str, Any]:
    return await browse_opcua_nodes(request)


@app.post("/api/database/test")
def database_test(connection: DatabaseConnection) -> dict[str, Any]:
    if connection.type == "sqlite":
        path = resolve_path(connection.path)
        if not path.exists():
            raise HTTPException(status_code=404, detail=f"Database not found: {path}")
        with sqlite_connect(path) as conn:
            conn.execute("SELECT 1").fetchone()
        return {"ok": True, "message": "SQLite connection succeeded.", "path": str(path)}

    conn = mysql_connect(connection)
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT 1 AS ok")
            row = cursor.fetchone()
    finally:
        conn.close()
    return {"ok": True, "message": "MySQL connection succeeded.", "result": row, "database": connection.database}


@app.post("/api/database/schema")
def database_schema(connection: DatabaseConnection) -> dict[str, Any]:
    return inspect_database_schema(connection)


@app.post("/api/database/query-preview")
def database_query_preview(request: QueryPreviewRequest) -> dict[str, Any]:
    return run_query(request)


@app.get("/api/data-sources")
def list_data_sources() -> list[dict[str, Any]]:
    with sqlite_connect(APP_DB) as conn:
        rows = conn.execute("SELECT * FROM data_sources ORDER BY id DESC").fetchall()
    return [row_to_data_source(row) for row in rows]


@app.post("/api/data-sources")
def create_data_source(payload: dict[str, Any]) -> dict[str, Any]:
    name = payload.get("name") or "Untitled Data Source"
    kind = payload.get("kind") or "database"
    config = payload.get("config") or {}
    with sqlite_connect(APP_DB) as conn:
        cursor = conn.execute(
            "INSERT INTO data_sources (name, kind, config_json, created_at, updated_at) VALUES (?, ?, ?, ?, ?)",
            (name, kind, json.dumps(config, ensure_ascii=False), now_iso(), now_iso()),
        )
    return {"id": cursor.lastrowid, "name": name, "kind": kind, "config": config}


@app.get("/api/opcua/points")
def list_opcua_points() -> list[dict[str, Any]]:
    with sqlite_connect(APP_DB) as conn:
        rows = conn.execute("SELECT * FROM data_sources WHERE kind = 'opcua_point' ORDER BY name").fetchall()
    return [row_to_data_source(row) for row in rows]


@app.post("/api/opcua/points")
def create_opcua_point(payload: OpcUaPointPayload) -> dict[str, Any]:
    alias = payload.alias.strip() or payload.display_name or payload.browse_name or payload.node_id
    config = payload.model_dump()
    config["alias"] = alias
    with sqlite_connect(APP_DB) as conn:
        row = conn.execute(
            "SELECT id FROM data_sources WHERE kind = 'opcua_point' AND json_extract(config_json, '$.node_id') = ?",
            (payload.node_id,),
        ).fetchone()
        if row:
            conn.execute(
                "UPDATE data_sources SET name = ?, config_json = ?, updated_at = ? WHERE id = ?",
                (alias, json.dumps(config, ensure_ascii=False), now_iso(), row["id"]),
            )
            point_id = row["id"]
        else:
            cursor = conn.execute(
                "INSERT INTO data_sources (name, kind, config_json, created_at, updated_at) VALUES (?, ?, ?, ?, ?)",
                (alias, "opcua_point", json.dumps(config, ensure_ascii=False), now_iso(), now_iso()),
            )
            point_id = cursor.lastrowid
    return {"id": point_id, "name": alias, "kind": "opcua_point", "config": config}


@app.put("/api/opcua/points/{point_id}")
def update_opcua_point(point_id: int, payload: OpcUaPointPayload) -> dict[str, Any]:
    alias = payload.alias.strip() or payload.display_name or payload.browse_name or payload.node_id
    config = payload.model_dump()
    config["alias"] = alias
    with sqlite_connect(APP_DB) as conn:
        row = conn.execute("SELECT id FROM data_sources WHERE id = ? AND kind = 'opcua_point'", (point_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="OPC UA point not found.")
        conn.execute(
            "UPDATE data_sources SET name = ?, config_json = ?, updated_at = ? WHERE id = ?",
            (alias, json.dumps(config, ensure_ascii=False), now_iso(), point_id),
        )
    return {"id": point_id, "name": alias, "kind": "opcua_point", "config": config}


@app.delete("/api/opcua/points/{point_id}")
def delete_opcua_point(point_id: int) -> dict[str, Any]:
    with sqlite_connect(APP_DB) as conn:
        conn.execute("DELETE FROM data_sources WHERE id = ? AND kind = 'opcua_point'", (point_id,))
    return {"ok": True}


@app.get("/api/report-templates")
def list_report_templates() -> list[dict[str, Any]]:
    with sqlite_connect(APP_DB) as conn:
        rows = conn.execute("SELECT id, name, config_json, created_at, updated_at FROM report_templates ORDER BY id DESC").fetchall()
    return [{**dict(row), "config": json.loads(row["config_json"])} for row in rows]


@app.get("/api/report-templates/{template_id}")
def read_report_template(template_id: int) -> dict[str, Any]:
    return get_template(template_id)


@app.post("/api/report-templates")
def create_report_template(payload: TemplatePayload) -> dict[str, Any]:
    config = payload.config
    config["name"] = payload.name
    with sqlite_connect(APP_DB) as conn:
        cursor = conn.execute(
            "INSERT INTO report_templates (name, config_json, created_at, updated_at) VALUES (?, ?, ?, ?)",
            (payload.name, json.dumps(config, ensure_ascii=False), now_iso(), now_iso()),
        )
    return {"id": cursor.lastrowid, "name": payload.name, "config": config}


@app.put("/api/report-templates/{template_id}")
def update_report_template(template_id: int, payload: TemplatePayload) -> dict[str, Any]:
    get_template(template_id)
    config = payload.config
    config["name"] = payload.name
    with sqlite_connect(APP_DB) as conn:
        conn.execute(
            "UPDATE report_templates SET name = ?, config_json = ?, updated_at = ? WHERE id = ?",
            (payload.name, json.dumps(config, ensure_ascii=False), now_iso(), template_id),
        )
    return {"id": template_id, "name": payload.name, "config": config}


@app.post("/api/report-templates/{template_id}/copy")
def copy_report_template(template_id: int) -> dict[str, Any]:
    template = get_template(template_id)
    copied_name = f"{template['name']} 副本"
    template.pop("id", None)
    template["name"] = copied_name
    return create_report_template(TemplatePayload(name=copied_name, config=template))


@app.delete("/api/report-templates/{template_id}")
def delete_report_template(template_id: int) -> dict[str, Any]:
    get_template(template_id)
    with sqlite_connect(APP_DB) as conn:
        conn.execute("DELETE FROM report_templates WHERE id = ?", (template_id,))
    return {"ok": True}


@app.post("/api/reports/generate")
async def reports_generate(request: GenerateReportRequest) -> dict[str, Any]:
    template = request.template or get_template(int(request.template_id or 0))
    try:
        report = await generate_report(template)
        if request.persist_run:
            with sqlite_connect(APP_DB) as conn:
                conn.execute(
                    "INSERT INTO report_runs (template_id, status, message, rendered_json, created_at) VALUES (?, ?, ?, ?, ?)",
                    (report.get("template_id"), "success", "ok", json.dumps(report, ensure_ascii=False), now_iso()),
                )
        return report
    except HTTPException as exc:
        if request.persist_run:
            with sqlite_connect(APP_DB) as conn:
                conn.execute(
                    "INSERT INTO report_runs (template_id, status, message, rendered_json, created_at) VALUES (?, ?, ?, ?, ?)",
                    (template.get("id"), "failed", str(exc.detail), None, now_iso()),
                )
        raise


@app.post("/api/reports/export/html")
async def reports_export_html(request: GenerateReportRequest) -> StreamingResponse:
    template = request.template or get_template(int(request.template_id or 0))
    report = await generate_report(template)
    content = report_to_html(report).encode("utf-8")
    return StreamingResponse(io.BytesIO(content), media_type="text/html; charset=utf-8", headers={"Content-Disposition": 'attachment; filename="report.html"'})


@app.post("/api/reports/export/excel")
async def reports_export_excel(request: GenerateReportRequest) -> StreamingResponse:
    template = request.template or get_template(int(request.template_id or 0))
    report = await generate_report(template)
    content = make_excel(report)
    return StreamingResponse(io.BytesIO(content), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="report.xlsx"'})


@app.post("/api/reports/export/pdf")
async def reports_export_pdf(request: GenerateReportRequest) -> StreamingResponse:
    template = request.template or get_template(int(request.template_id or 0))
    report = await generate_report(template)
    content = make_pdf(report)
    return StreamingResponse(io.BytesIO(content), media_type="application/pdf", headers={"Content-Disposition": 'attachment; filename="report.pdf"'})


@app.get("/api/report-runs")
def list_report_runs() -> list[dict[str, Any]]:
    with sqlite_connect(APP_DB) as conn:
        rows = conn.execute("SELECT id, template_id, status, message, created_at FROM report_runs ORDER BY id DESC LIMIT 50").fetchall()
    return [dict(row) for row in rows]
